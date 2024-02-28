import openpyxl
from openpyxl.styles import PatternFill
import datetime

# path = 'Шаблон.xlsx'
#
# work_book = openpyxl.load_workbook(path)
# sheet = work_book.active


class Workname():
    def __init__(self, names, data, lines = "ABCDEFGHIJKLMNOPQRSTUVWXYZ", smena = 1, hour = 8):
        self.lines = lines
        self.smena = smena  # 2, 3
        self.hour = hour  # 10, 12
        self.fun(names,data)
        self.counter_rows = 4
        path = 'Шаблон.xlsx'

        self.work_book = openpyxl.load_workbook(path)
        self.sheet = self.work_book.active

    def fun(self, names, data):
        self.names = names
        self.example_dict = dict()
        for i in data:
            self.example_dict[i[0]] = dict(zip(self.names, i))
        print(self.example_dict)

    def calc(self, example_dict):
        day = (float(example_dict["Объем работ"])*float(example_dict["Норма времени, чел.-ч."]))/(int(example_dict["Кол-во персонала в смену"])*int(self.ed_izm(example_dict["Единицы измерения"]))*self.smena*self.hour)
        if day*100%100 >= 20:
            day = int(day)+1
        else:
            day = int(day)
        return day


    def ed_izm(self, data):
        digit = data.split(" ")
        if len(digit) > 1:
            return digit[0]
        return 1

    def draw(self, example_dict, start_date):
        start_date = datetime.datetime.strptime(start_date, "%d.%m.%Y")
        self.sheet[f"A{self.counter_rows}"] = example_dict["Наименование работ"]
        duration = self.calc(example_dict)
        end_date = start_date + datetime.timedelta(days=duration)
        self.sheet[f"B{self.counter_rows}"] = duration
        self.sheet[f"C{self.counter_rows}"] = start_date.strftime("%d.%m.%Y")
        self.sheet[f"D{self.counter_rows}"] = end_date.strftime("%d.%m.%Y")
        fill = PatternFill(start_color='4F8E80', end_color='4F8E80', fill_type='solid')
        self.sheet["G4"].fill = fill
        for i in range(duration):
            self.sheet[f"{self.lines[6 + i]}3"] = (start_date + datetime.timedelta(days=i)).strftime("%d.%m")
            self.sheet[f"{self.lines[6+i]}{self.counter_rows}"].fill = fill


    def save(self):
        self.work_book.save('text.xlsx')
    def close(self):
        self.work_book.close()


if __name__ == "__main__":
    names = "Наименование работ\tОбоснование норм\tКол-во персонала/машин в смену\tЕд. изм.\tНорма затрат труда, чел.-ч.\tОбъем V"
    example_1 = "Монтаж плит\t30-018-01-1\t6\t1\t2.76\t40.6"
    example_2 = "Устройство обмазочной гидроизоляции\tЕ4-3-184\t4\t1 м3\t0.27\t78"
    example_3 = "Обсыпка трубы\tЕ2-1-34 5\t100 м3\t0.49\t2.1"
    names = names.split("\t")
    data = [example_1.split("\t"), example_2.split("\t"), example_3.split("\t")]
    work1 = Workname(names,data)
    print(work1.calc(work1.example_dict["Монтаж плит"]))
    work1.draw(work1.example_dict["Монтаж плит"], "16.09.2023")
    work1.save()
    work1.close()
